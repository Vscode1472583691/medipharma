"""
Excel Export Module
Handles exporting data to Excel sheets with year & month wise organization
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from io import BytesIO
from collections import defaultdict


def apply_header_formatting(sheet, headers):
    """Apply formatting to header row"""
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    
    for col_num, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def auto_fit_columns(sheet):
    """Auto-fit column widths"""
    for column in sheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        sheet.column_dimensions[column_letter].width = adjusted_width


def get_gst_split(item):
    cgst = getattr(item, 'cgst_percent', None)
    sgst = getattr(item, 'sgst_percent', None)
    if cgst is None or sgst is None:
        total = getattr(item, 'gst_percent', 0) or 0
        return total / 2.0, total / 2.0
    return cgst, sgst


def export_purchases(purchases, shop_name=""):
    """Export purchases data to Excel"""
    wb = Workbook()
    
    # Group purchases by month-year
    purchases_by_month = defaultdict(list)
    for purchase in purchases:
        month_key = purchase.date.strftime("%Y-%m") if purchase.date else "Unknown"
        purchases_by_month[month_key].append(purchase)
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Create sheet for each month
    for month_key in sorted(purchases_by_month.keys()):
        month_purchases = purchases_by_month[month_key]
        sheet_name = month_key.replace("-", "_")[:31]  # Excel sheet name limit
        ws = wb.create_sheet(title=sheet_name)
        
        headers = ["ID", "Date", "Supplier", "Product", "Batch No", "MRP", "Quantity", "Unit Price", "CGST %", "SGST %", "Discount", "Total"]
        apply_header_formatting(ws, headers)
        
        row = 2
        total_amount = 0
        for purchase in month_purchases:
            supplier = purchase.__dict__.get('supplier')
            supplier_name = supplier.name if supplier else "Unknown"
            product = purchase.__dict__.get('product')
            product_name = product.name if product else "Unknown"
            
            ws.cell(row=row, column=1).value = purchase.id
            ws.cell(row=row, column=2).value = purchase.date
            ws.cell(row=row, column=3).value = supplier_name
            ws.cell(row=row, column=4).value = product_name
            ws.cell(row=row, column=5).value = getattr(purchase, 'sku', None)
            ws.cell(row=row, column=6).value = purchase.mrp_price
            ws.cell(row=row, column=7).value = purchase.qty
            ws.cell(row=row, column=8).value = purchase.price
            cgst, sgst = get_gst_split(purchase)
            ws.cell(row=row, column=9).value = cgst
            ws.cell(row=row, column=10).value = sgst
            ws.cell(row=row, column=11).value = purchase.discount
            ws.cell(row=row, column=12).value = purchase.total
            
            total_amount += purchase.total
            row += 1
        
        # Add total row
        ws.cell(row=row, column=11).value = "TOTAL:"
        ws.cell(row=row, column=11).font = Font(bold=True)
        ws.cell(row=row, column=12).value = total_amount
        ws.cell(row=row, column=12).font = Font(bold=True)
        
        auto_fit_columns(ws)
    
    # Create summary sheet if multiple months
    if len(purchases_by_month) > 1:
        summary_ws = wb.create_sheet("Summary", 0)
        summary_ws["A1"] = "Purchase Summary"
        summary_ws["A1"].font = Font(bold=True, size=14)
        
        summary_ws["A3"] = "Month"
        summary_ws["B3"] = "Total Amount"
        apply_header_formatting(summary_ws, ["Month", "Total Amount"])
        
        row = 4
        for month_key in sorted(purchases_by_month.keys()):
            total = sum(p.total for p in purchases_by_month[month_key])
            summary_ws.cell(row=row, column=1).value = month_key
            summary_ws.cell(row=row, column=2).value = total
            row += 1
        
        auto_fit_columns(summary_ws)
    
    return wb


def export_purchase_returns(returns, shop_name=""):
    wb = Workbook()
    returns_by_month = defaultdict(list)
    for item in returns:
        month_key = item.date.strftime("%Y-%m") if item.date else "Unknown"
        returns_by_month[month_key].append(item)
    wb.remove(wb.active)

    for month_key in sorted(returns_by_month.keys()):
        month_returns = returns_by_month[month_key]
        sheet_name = month_key.replace("-", "_")[:31]
        ws = wb.create_sheet(title=sheet_name)
        headers = ["ID", "Date", "Supplier", "Product", "Batch No", "Quantity", "Unit Price", "Total", "Remark"]
        apply_header_formatting(ws, headers)
        row = 2
        total_amount = 0
        for item in month_returns:
            supplier = item.__dict__.get('supplier')
            supplier_name = supplier.name if supplier else "Unknown"
            product = item.__dict__.get('product')
            product_name = product.name if product else "Unknown"
            ws.cell(row=row, column=1).value = item.id
            ws.cell(row=row, column=2).value = item.date
            ws.cell(row=row, column=3).value = supplier_name
            ws.cell(row=row, column=4).value = product_name
            ws.cell(row=row, column=5).value = getattr(item, 'sku', None)
            ws.cell(row=row, column=6).value = item.qty
            ws.cell(row=row, column=7).value = item.price
            ws.cell(row=row, column=8).value = item.total
            ws.cell(row=row, column=9).value = item.remark or ""
            total_amount += item.total
            row += 1
        ws.cell(row=row, column=7).value = "TOTAL:"
        ws.cell(row=row, column=7).font = Font(bold=True)
        ws.cell(row=row, column=8).value = total_amount
        ws.cell(row=row, column=8).font = Font(bold=True)
        auto_fit_columns(ws)

    return wb


def export_sale_returns(returns, shop_name=""):
    wb = Workbook()
    returns_by_month = defaultdict(list)
    for item in returns:
        month_key = item.date.strftime("%Y-%m") if item.date else "Unknown"
        returns_by_month[month_key].append(item)
    wb.remove(wb.active)

    for month_key in sorted(returns_by_month.keys()):
        month_returns = returns_by_month[month_key]
        sheet_name = month_key.replace("-", "_")[:31]
        ws = wb.create_sheet(title=sheet_name)
        headers = ["ID", "Date", "Customer", "Product", "Batch No", "Quantity", "Unit Price", "Total", "Remark"]
        apply_header_formatting(ws, headers)
        row = 2
        total_amount = 0
        for item in month_returns:
            customer = item.__dict__.get('customer')
            customer_name = customer.name if customer else "Unknown"
            product = item.__dict__.get('product')
            product_name = product.name if product else "Unknown"
            ws.cell(row=row, column=1).value = item.id
            ws.cell(row=row, column=2).value = item.date
            ws.cell(row=row, column=3).value = customer_name
            ws.cell(row=row, column=4).value = product_name
            ws.cell(row=row, column=5).value = getattr(item, 'sku', None)
            ws.cell(row=row, column=6).value = item.qty
            ws.cell(row=row, column=7).value = item.price
            ws.cell(row=row, column=8).value = item.total
            ws.cell(row=row, column=9).value = item.remark or ""
            total_amount += item.total
            row += 1
        ws.cell(row=row, column=7).value = "TOTAL:"
        ws.cell(row=row, column=7).font = Font(bold=True)
        ws.cell(row=row, column=8).value = total_amount
        ws.cell(row=row, column=8).font = Font(bold=True)
        auto_fit_columns(ws)

    return wb


def export_sales(sales, shop_name=""):
    """Export sales data to Excel"""
    wb = Workbook()
    
    # Group sales by month-year
    sales_by_month = defaultdict(list)
    for sale in sales:
        month_key = sale.date.strftime("%Y-%m") if sale.date else "Unknown"
        sales_by_month[month_key].append(sale)
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Create sheet for each month
    for month_key in sorted(sales_by_month.keys()):
        month_sales = sales_by_month[month_key]
        sheet_name = month_key.replace("-", "_")[:31]
        ws = wb.create_sheet(title=sheet_name)
        
        headers = ["ID", "Date", "Customer", "Product", "Batch No", "MRP", "Quantity", "Unit Price", "CGST %", "SGST %", "Discount", "Total", "Seller"]
        apply_header_formatting(ws, headers)
        
        row = 2
        total_amount = 0
        for sale in month_sales:
            customer = sale.__dict__.get('customer')
            customer_name = customer.name if customer else "Unknown"
            product = sale.__dict__.get('product')
            product_name = product.name if product else "Unknown"
            product_sku = getattr(sale, 'sku', None) or (product.sku if product else "Unknown")
            product_mrp = sale.mrp_price or (product.mrp_price if product else 0)
            
            ws.cell(row=row, column=1).value = sale.id
            ws.cell(row=row, column=2).value = sale.date
            ws.cell(row=row, column=3).value = customer_name
            ws.cell(row=row, column=4).value = product_name
            ws.cell(row=row, column=5).value = product_sku
            ws.cell(row=row, column=6).value = product_mrp
            ws.cell(row=row, column=7).value = sale.qty
            ws.cell(row=row, column=8).value = sale.price
            cgst, sgst = get_gst_split(sale)
            ws.cell(row=row, column=9).value = cgst
            ws.cell(row=row, column=10).value = sgst
            ws.cell(row=row, column=11).value = sale.discount
            ws.cell(row=row, column=12).value = sale.total
            ws.cell(row=row, column=13).value = sale.seller_name or ""
            
            total_amount += sale.total
            row += 1
        
        # Add total row
        ws.cell(row=row, column=12).value = "TOTAL:"
        ws.cell(row=row, column=12).font = Font(bold=True)
        ws.cell(row=row, column=13).value = total_amount
        ws.cell(row=row, column=13).font = Font(bold=True)
        
        auto_fit_columns(ws)
    
    # Create summary sheet if multiple months
    if len(sales_by_month) > 1:
        summary_ws = wb.create_sheet("Summary", 0)
        summary_ws["A1"] = "Sales Summary"
        summary_ws["A1"].font = Font(bold=True, size=14)
        
        summary_ws["A3"] = "Month"
        summary_ws["B3"] = "Total Amount"
        apply_header_formatting(summary_ws, ["Month", "Total Amount"])
        
        row = 4
        for month_key in sorted(sales_by_month.keys()):
            total = sum(s.total for s in sales_by_month[month_key])
            summary_ws.cell(row=row, column=1).value = month_key
            summary_ws.cell(row=row, column=2).value = total
            row += 1
        
        auto_fit_columns(summary_ws)
    
    return wb


def export_products(products):
    """Export products/inventory data to Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventory"
    
    headers = ["ID", "Product Name", "Batch No", "MRP", "Quantity", "Selling Price", "Purchase Price", "GST %", "Expiry Date"]
    apply_header_formatting(ws, headers)
    
    row = 2
    for product in products:
        ws.cell(row=row, column=1).value = product.id
        ws.cell(row=row, column=2).value = product.name
        ws.cell(row=row, column=3).value = product.sku
        ws.cell(row=row, column=4).value = product.mrp_price
        ws.cell(row=row, column=5).value = product.quantity
        ws.cell(row=row, column=6).value = product.price
        ws.cell(row=row, column=7).value = product.purchase_price
        ws.cell(row=row, column=8).value = product.gst_percent
        ws.cell(row=row, column=9).value = product.expiry_date
        row += 1
    
    auto_fit_columns(ws)
    return wb


def export_suppliers(suppliers):
    """Export suppliers data to Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Suppliers"
    
    headers = ["ID", "Supplier Name", "Phone", "GST Number"]
    apply_header_formatting(ws, headers)
    
    row = 2
    for supplier in suppliers:
        ws.cell(row=row, column=1).value = supplier.id
        ws.cell(row=row, column=2).value = supplier.name
        ws.cell(row=row, column=3).value = supplier.phone or ""
        ws.cell(row=row, column=4).value = supplier.gst or ""
        row += 1
    
    auto_fit_columns(ws)
    return wb


def export_customers(customers):
    """Export customers data to Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Customers"
    
    headers = ["ID", "Customer Name", "Phone", "Address", "Payment Due", "Payment Complete", "Remarks"]
    apply_header_formatting(ws, headers)
    
    row = 2
    for customer in customers:
        ws.cell(row=row, column=1).value = customer.id
        ws.cell(row=row, column=2).value = customer.name
        ws.cell(row=row, column=3).value = customer.phone or ""
        ws.cell(row=row, column=4).value = customer.address or ""
        ws.cell(row=row, column=5).value = customer.payment_due
        ws.cell(row=row, column=6).value = "Yes" if customer.payment_complete else "No"
        ws.cell(row=row, column=7).value = customer.remark or ""
        row += 1
    
    auto_fit_columns(ws)
    return wb


def export_payments(payments):
    """Export payments data to Excel"""
    wb = Workbook()
    
    # Group payments by month-year
    payments_by_month = defaultdict(list)
    for payment in payments:
        month_key = payment.date.strftime("%Y-%m") if payment.date else "Unknown"
        payments_by_month[month_key].append(payment)
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Create sheet for each month
    for month_key in sorted(payments_by_month.keys()):
        month_payments = payments_by_month[month_key]
        sheet_name = month_key.replace("-", "_")[:31]
        ws = wb.create_sheet(title=sheet_name)
        
        headers = ["ID", "Date", "Party Type", "Party ID", "Amount", "Mode", "Payment Method", "Direction"]
        apply_header_formatting(ws, headers)
        
        row = 2
        total_amount = 0
        for payment in month_payments:
            ws.cell(row=row, column=1).value = payment.id
            ws.cell(row=row, column=2).value = payment.date
            ws.cell(row=row, column=3).value = payment.party_type or ""
            ws.cell(row=row, column=4).value = payment.party_id
            ws.cell(row=row, column=5).value = payment.amount
            ws.cell(row=row, column=6).value = payment.mode or ""
            ws.cell(row=row, column=7).value = payment.payment_method or ""
            ws.cell(row=row, column=8).value = payment.direction or ""
            
            total_amount += payment.amount
            row += 1
        
        # Add total row
        ws.cell(row=row, column=4).value = "TOTAL:"
        ws.cell(row=row, column=4).font = Font(bold=True)
        ws.cell(row=row, column=5).value = total_amount
        ws.cell(row=row, column=5).font = Font(bold=True)
        
        auto_fit_columns(ws)
    
    # Create summary sheet if multiple months
    if len(payments_by_month) > 1:
        summary_ws = wb.create_sheet("Summary", 0)
        summary_ws["A1"] = "Payments Summary"
        summary_ws["A1"].font = Font(bold=True, size=14)
        
        summary_ws["A3"] = "Month"
        summary_ws["B3"] = "Total Amount"
        apply_header_formatting(summary_ws, ["Month", "Total Amount"])
        
        row = 4
        for month_key in sorted(payments_by_month.keys()):
            total = sum(p.amount for p in payments_by_month[month_key])
            summary_ws.cell(row=row, column=1).value = month_key
            summary_ws.cell(row=row, column=2).value = total
            row += 1
        
        auto_fit_columns(summary_ws)
    
    return wb


def export_bills(sales, shop_info=None):
    """Export bills/invoices data to Excel"""
    wb = Workbook()
    
    # Group bills by month-year
    bills_by_month = defaultdict(list)
    for sale in sales:
        month_key = sale.date.strftime("%Y-%m") if sale.date else "Unknown"
        bills_by_month[month_key].append(sale)
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Create sheet for each month
    for month_key in sorted(bills_by_month.keys()):
        month_bills = bills_by_month[month_key]
        sheet_name = month_key.replace("-", "_")[:31]
        ws = wb.create_sheet(title=sheet_name)
        
        headers = ["Bill ID", "Date", "Customer", "Product", "MRP", "Qty", "Unit Price", "CGST %", "SGST %", "Discount", "Total", "Seller"]
        apply_header_formatting(ws, headers)
        
        row = 2
        total_amount = 0
        for bill in month_bills:
            customer = bill.__dict__.get('customer')
            customer_name = customer.name if customer else "Unknown"
            product = bill.__dict__.get('product')
            product_name = product.name if product else "Unknown"
            
            ws.cell(row=row, column=1).value = bill.id
            ws.cell(row=row, column=2).value = bill.date
            ws.cell(row=row, column=3).value = customer_name
            ws.cell(row=row, column=4).value = product_name
            ws.cell(row=row, column=5).value = bill.mrp_price
            ws.cell(row=row, column=6).value = bill.qty
            ws.cell(row=row, column=7).value = bill.price
            ws.cell(row=row, column=8).value = bill.cgst_percent
            ws.cell(row=row, column=9).value = bill.sgst_percent
            ws.cell(row=row, column=10).value = bill.discount
            ws.cell(row=row, column=11).value = bill.total
            ws.cell(row=row, column=12).value = bill.seller_name or ""
            
            total_amount += bill.total
            row += 1
        
        # Add total row
        ws.cell(row=row, column=12).value = "TOTAL:"
        ws.cell(row=row, column=12).font = Font(bold=True)
        ws.cell(row=row, column=13).value = total_amount
        ws.cell(row=row, column=13).font = Font(bold=True)
        
        auto_fit_columns(ws)
    
    # Create summary sheet if multiple months
    if len(bills_by_month) > 1:
        summary_ws = wb.create_sheet("Summary", 0)
        summary_ws["A1"] = "Bills Summary"
        summary_ws["A1"].font = Font(bold=True, size=14)
        
        summary_ws["A3"] = "Month"
        summary_ws["B3"] = "Total Amount"
        apply_header_formatting(summary_ws, ["Month", "Total Amount"])
        
        row = 4
        for month_key in sorted(bills_by_month.keys()):
            total = sum(b.total for b in bills_by_month[month_key])
            summary_ws.cell(row=row, column=1).value = month_key
            summary_ws.cell(row=row, column=2).value = total
            row += 1
        
        auto_fit_columns(summary_ws)
    
    return wb


def excel_to_bytes(wb):
    """Convert Excel workbook to bytes for download"""
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()
